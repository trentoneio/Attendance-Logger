# -*- coding: utf-8 -*-
"""
Created on Tue Aug  9 09:13:52 2022

@author: Trenton Foster

"""


import tkinter as tk
from tkinter import messagebox
import ctypes
import openpyxl
import sys
import os




global counter
global workbook
counter = 0


#Allow users to select file to be processed
def openfilepath(string):
    root = tk.Tk()
    root.withdraw() # we don't want a full GUI, so keep the root window from appearing
    from tkinter import filedialog
    file_selected = filedialog.askopenfilename(title = string)
    root.destroy()
    return file_selected


def updateexceldoc(studentnum,master,worksheet):
    global counter
    global workbook
    counter += 1
    worksheet = workbook.active
    cell = worksheet.cell(row = counter, column = 1)
    cell.value = studentnum

    if studentnum in master:
        #Mbox("","Student is paid and registered\nYES YES YES",0)
        messagebox.showinfo("","Student is paid and registered\nYES YES YES")
        tempcell = worksheet.cell(row = counter, column = 2)
        tempcell.value = "paid"

    else:
        #Mbox("","Student is NOT paid and registered\n\nNO NO NO",0)
        messagebox.showinfo("","Student is NOT paid and registered\n\nNO NO NO")
        tempcell = worksheet.cell(row = counter, column = 2)
        tempcell.value = "NOT paid"




def quitloop():
    global workbook
    #Mbox("","The next box will request a file save location",0)
    messagebox.showinfo("","The next box will request a file save location")
    try:
        workbook.save(openfilepath("Select path to save file to"))
        #Mbox("Finished!","Document saved to given folder path (or master path if none given).",0)
        messagebox.showinfo("Finished!","Document saved to given folder path (or master path if none given). You will now need to FORCE EXIT the program because Mac sucks.")
        root.destroy()
        return()

    except:
        head_tail = os.path.split(master_file)
        save_file = head_tail[0]+"/TellTrentThankYou.xlsx"
        workbook.save(save_file)
        messagebox.showinfo("Finished!","Document saved to given folder path (or master path if none given). You will now need to FORCE EXIT the program because Mac sucks.")
        root.destroy()
        return()



root = tk.Tk().withdraw()
#Mbox("", "Before pressing OK, ensure that you have the excel file containing the student numbers of all currently paid and registered members.\nIn the next window, you will be asked to select that excel file.",0)
messagebox.showinfo('', "Before pressing OK, ensure that you have the excel file containing the student numbers of all currently paid and registered members.\nIn the next window, you will be asked to select that excel file.")
master_file = openfilepath("Select master file (file with student IDs)")

try:
    wb = openpyxl.load_workbook(filename= master_file)
except:
    #Mbox("FATAL ERROR", "Incorrect file type selected. File must be a .xlsx excel file.\nRestart program and try again.",0)
    messagebox.showinfo("","Incorrect file type selected. File must be a .xlsx excel file.\nRestart program and try again.")
    sys.exit()




sheet = wb.active
rowrange = sheet.max_row
Masterlist = []
for i in range (1, rowrange + 1):
    cell = sheet.cell(row = i, column = 1)
    Masterlist.append(str(cell.value))




#finding part number and test name for data
root = tk.Tk()
root.wm_title("Enter student nubmers")
root.geometry('500x75')
student_var = tk.Entry(root) #initialize text box for student number
student_var.insert(tk.END,'16281689')
workbook = openpyxl.Workbook()
worksheet = workbook.active


updatebutton = tk.Button(root, text = "Update", command= lambda: updateexceldoc(student_var.get(),Masterlist,worksheet))
finalbutton = tk.Button(root, text = "Finish", command = lambda: quitloop())



#packing tk window. Started by packing onto the leftmost part of the window, and then stacking components on top of that
tk.Label(root, text="Enter Student Number: ").pack(side=tk.LEFT)
student_var.pack(side=tk.LEFT)
updatebutton.pack(side=tk.LEFT)
finalbutton.pack(side=tk.BOTTOM)

root.mainloop()
