# -*- coding: utf-8 -*-
"""
Created on Tue Aug  9 09:13:52 2022

@author: Trenton Foster

"""


import tkinter as tk
import ctypes
import openpyxl
import sys
import os




global counter
global workbook
counter = 0


#Allow users to select file to be processed
def openfilepath(string):
    root = tk.Tk()
    root.withdraw() # we don't want a full GUI, so keep the root window from appearing
    from tkinter import filedialog
    file_selected = filedialog.askopenfilename(title = string)
    root.destroy()
    return file_selected


def Mbox(title, text, style):
    return ctypes.windll.user32.MessageBoxW(0, text, title, style)


def updateexceldoc(studentnum,master,worksheet):
    global counter
    global workbook
    counter += 1
    worksheet = workbook.active
    cell = worksheet.cell(row = counter, column = 1)
    cell.value = studentnum

    if studentnum in master:
        Mbox("","Student is paid and registered\nYES YES YES",0)
        tempcell = worksheet.cell(row = counter, column = 2)
        tempcell.value = "paid"

    else:
        Mbox("","Student is NOT paid and registered\n\nNO NO NO",0)
        tempcell = worksheet.cell(row = counter, column = 2)
        tempcell.value = "NOT paid"




def quitloop():
    global workbook
    Mbox("","The next box will request a file save location",0)
    try:
        workbook.save(openfilepath("Select path to save file to"))
    except:
        head_tail = os.path.split(master_file)
        save_file = head_tail[0]+"/TellTrentThankYou.xlsx"
        workbook.save(save_file)
    root.destroy()




Mbox("", "Before pressing OK, ensure that you have the excel file containing the student numbers of all currently paid and registered members.\nIn the next window, you will be asked to select that excel file.",0)
master_file = openfilepath("Select master file (file with student IDs)")

try:
    wb = openpyxl.load_workbook(filename= master_file)
except:
    Mbox("FATAL ERROR", "Incorrect file type selected. File must be a .xlsx excel file.\nRestart program and try again.",0)
    sys.exit()




sheet = wb.active
rowrange = sheet.max_row
Masterlist = []
for i in range (1, rowrange + 1):
    cell = sheet.cell(row = i, column = 1)
    Masterlist.append(str(cell.value))




#finding part number and test name for data
root = tk.Tk()
root.wm_title("Enter student nubmers")
root.geometry('400x75')
student_var = tk.Entry() #initialize text box for student number
student_var.insert(tk.END,'16281689')
workbook = openpyxl.Workbook()
worksheet = workbook.active


updatebutton = tk.Button(root, text = "Update", command= lambda: updateexceldoc(student_var.get(),Masterlist,worksheet))
finalbutton = tk.Button(root, text = "Finish", command = lambda: quitloop())



#packing tk window. Started by packing onto the leftmost part of the window, and then stacking components on top of that
tk.Label(root, text="Enter Student Number: ").grid(row=1, column=1)
student_var.grid(row=1, column=2)
updatebutton.grid(row=1, column=3)
finalbutton.grid(row=3, column=3)

root.mainloop()

Mbox("Finished!","Document saved to given folder path (or master path if none given).",0)
